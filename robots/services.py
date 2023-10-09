from datetime import datetime, timedelta
from django.db.models import Count
from django.http import HttpResponse
from openpyxl import Workbook

from robots.models import Robot


def robots_last_week(request):
    workbook = Workbook()
    index = 0
    time_now = datetime.now()
    time_week_ago = time_now - timedelta(days=7)
    robots_models = Robot.objects.filter(created__range=(time_week_ago, time_now)
                                         ).values_list('model', flat=True).distinct()
    for model in robots_models:
        robots = Robot.objects.filter(model=model, created__range=(time_week_ago, time_now)
                                      ).values('model', 'version', 'serial').annotate(count=Count('serial'))
        sheet = workbook.create_sheet(title=model, index=index)
        index += 1
        sheet.cell(row=1, column=1, value='Модель')
        sheet.cell(row=1, column=2, value='Версия')
        sheet.cell(row=1, column=3, value='Количество за неделю')
        row_num = 2
        for robot in robots:
            sheet.cell(row=row_num, column=1, value=robot['model'])
            sheet.cell(row=row_num, column=2, value=robot['version'])
            sheet.cell(row=row_num, column=3, value=robot['count'])
            row_num += 1
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=robots.xlsx'

    workbook.save(response)

    return response
